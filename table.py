from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment
campos = ['FECHA', 'HORA', 'CATEGORIA', 'ALARMA', 'ESTADO']
col = ['A', 'B', 'C', 'D', 'E']
col_width = [12, 10, 12, 52, 15]


def create_table(name):
    wb = Workbook()
    ws = wb.active
    ws.title = name
    for i in range(len(campos)):
        ws.column_dimensions[col[i]].width = col_width[i]
        celda = col[i] + '1'
        ws.cell(1, i + 1, campos[i])
        ws[celda].fill = PatternFill(fgColor="0070C0", fill_type="solid")
        ws[celda].font = Font(color="FFFFFF")
        ws[celda].alignment = Alignment(horizontal='center')
    return wb


def load_table(wb, fila, log):
    ws = wb.active
    for i in range(1, len(log) + 1):
        celda = col[i - 1] + str(fila)
        ws.cell(fila, i, log[i - 1])
        ws[celda].alignment = Alignment(horizontal='center')
        if (fila % 2) == 0:
            ws[celda].fill = PatternFill(fgColor="EEECE1", fill_type="solid")
        else:
            ws[celda].fill = PatternFill(fgColor="BFBFBF", fill_type="solid")

